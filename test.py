import telebot
from telebot import types
import openpyxl
import random
from transliterate import translit
import datetime
import requests

# Telegram Bot
bot = telebot.TeleBot('5312004578:AAFfG4l8EQLf4TXj7vn9OZo56CGGG4EFKK8')


stickers_id = []
operators_id = [198384562, 829591706, 523811214]
admins_id = [198384562, 829591706, 523811214]
monday_menu = {}
tuesday_menu = {}
wednesday_menu = {}
thursday_menu = {}
friday_menu = {}
bill_list = []
days_suff = {'Понедельник': 'понедельник', 'Вторник': 'вторник', 'Среда': 'среду', 'Четверг': 'четверг', 'Пятница': 'пятницу'}
excel_days_cells = {'monday': (3, 6, 4), 'tuesday': (10, 13, 11), 'wednesday': (17, 20, 18), 'thursday': (24, 27, 25), 'friday': (31, 34, 32)}
days_with_menu = [monday_menu, tuesday_menu, wednesday_menu, thursday_menu, friday_menu]
days_for_sort = {'Понедельник': 1, 'Вторник': 2, 'Среда': 3, 'Четверг': 4, 'Пятница': 5}
just_days = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница']
emojis = {'суп': '🍲', 'салат': '🥗', 'фалафель': '🧆', 'омлет': '🍳', 'каша': '🥣', 'бутерброд': '🥪', 'кекс': '🍮',
          'печенье': '🍪', 'картофель': '🍟', 'рис': '🍚', 'курица': '🍗', 'блины': '🥞', 'макароны': '🍝', 'рагу': '🥘',
          'панкейк': '🥞', 'цезарь': '🥗', 'борщ': '🍲', 'гречка': '🍛', 'компот': '🧃', 'кисель': '🧃', 'плов': '🍛'}


# Генератор ID заказа
def id_generator(name):
    numbers = '1234567890'
    digits = 'qwertyuioplkjhgfdsazxcvbnm'
    up_digits = 'QWERTYUIOPLKJHGFDSAZXCVBNM'
    all = [numbers, digits, up_digits]
    new_id = name[:3]
    while len(new_id) < 9:
        rand_list = random.choice(all)
        new_id += random.choice(rand_list)
    return(new_id)


# Транслит  имени
def translit_name(name):
    flag_name = False
    if name == None:
        return('_')
    for i in name:
        if i.lower() in 'йцукенгшщзхжъэдлорпавыфячсмитьбю':
            flag_name = True
            break
    if flag_name == True:
        first_name = translit(name, 'ru', reversed=True)
        return(first_name)
    else:
        return(name)


# Excel файл
file = openpyxl.open('Меню.xlsx')
file_list = file.active


# Создание меню
def create_menu(excel_day_cells):
    count = -1
    for j in excel_days_cells:
        count += 1
        for i in range(1, file_list.max_row + 1):
            name = file_list.cell(row=i, column=excel_days_cells[j][0]).value
            price = file_list.cell(row=i, column=excel_days_cells[j][1]).value
            if name == None or price == None:
                continue
            else:
                if len(name) > 40:
                    name = name[:40] + '...'
                else:
                    temp_name = name.split()
                    for q in temp_name:
                        if q.lower() in emojis:
                            name += ' ' + emojis[q.lower()]
                            break
                temp = price, i, excel_days_cells[j][2]
                days_with_menu[count].setdefault(name, temp)
    bot.send_message(198384562, 'Меню создано')
    print('Меню создано')
create_menu(excel_days_cells)


# Команда start
@bot.message_handler(commands=['start'])
@bot.message_handler(func=lambda msg: msg.text == 'Назад' or msg.text == 'Режим пользователя' or msg.text == 'Вернуться к выбору режима' or msg.text == 'Назад в выбор режима')
def start(message):
    if message.from_user.id not in admins_id or message.text == 'Режим пользователя':
        start_menu = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        menu_button = types.KeyboardButton(text='Меню')
        pay_button = types.KeyboardButton(text='Оплатить заказ')
        order_button = types.KeyboardButton(text='Мой заказ')
        admin_menu_button = types.KeyboardButton(text='Вернуться к выбору режима')
        if message.from_user.id in admins_id:
            start_menu.add(menu_button, order_button, pay_button, admin_menu_button)
        else:
            start_menu.add(menu_button, order_button, pay_button)
        bot.send_message(message.chat.id, 'Привет! Выбери один из пунктов меню ⬇', reply_markup=start_menu)
    elif message.from_user.id in admins_id or (message.from_user.id in admins_id and message.text == 'Режим администратора') or (message.from_user.id in admins_id and message.text == 'Вернуться к выбору режима') or (message.from_user.id in admins_id and message.text == 'Назад к выбору режима'):
        start_admin_menu = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        admin_button = types.KeyboardButton(text='Режим администратора')
        user_button = types.KeyboardButton(text='Режим пользователя')
        start_admin_menu.add(admin_button, user_button)
        bot.send_message(message.chat.id, 'Добро пожаловать, Босс 🌚', reply_markup=start_admin_menu)


# Режим администратора
@bot.message_handler(func=lambda msg: msg.text == 'Режим администратора')
def admin_console(message):
    admin_console = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    orders_button = types.KeyboardButton(text='Все заказы')
    pay_button = types.KeyboardButton(text='Чеки')
    back_admin_button = types.KeyboardButton(text='Назад в выбор режима')
    logs_button = types.KeyboardButton(text='Файл с логами')
    change_menu_button = types.KeyboardButton(text='Изменить меню на неделю')
    admin_console.add(orders_button, pay_button, logs_button,change_menu_button, back_admin_button)
    bot.send_message(message.chat.id, 'Что вы хотите, Босс? 🌝', reply_markup=admin_console)


# Изменить меню на неделю
@bot.message_handler(func=lambda msg: msg.text == 'Изменить меню на неделю')
def change_menu(message):
    bot.send_message(message.chat.id, 'Пришли excel файл с меню. Файл должен называться "Меню" и иметь только один лист')


# Новое меню
@bot.message_handler(func=lambda msg: msg.document.file_name == 'Меню.xlsx', content_types=['document'])
def get_document(message):
    file_id = message.document.file_id
    file_info = bot.get_file(file_id)
    file = requests.get('https://api.telegram.org/file/bot{0}/{1}'.format('5312004578:AAFfG4l8EQLf4TXj7vn9OZo56CGGG4EFKK8', file_info.file_path))
    with open('Меню.xlsx', 'wb') as f:
        f.write(file.content)
    with open('Заказы.txt', 'w') as f:
        f.write('test|test|test|130|test')
    with open('Логи.txt', 'w') as f:
        time = datetime.datetime.now()
        f.write(f'{str(time)[:19]} Загружено новое меню. Пользователь: {message.from_user.username}')
    bot.send_message(message.chat.id, 'Меню успешно изменено. Все прошлые заказы и логи были удалены')


# Все заказы
@bot.message_handler(func=lambda msg: msg.text == 'Все заказы')
def admin_orders(message):
    with open('Заказы.txt', 'r') as orders:
        orders_text = orders.readlines()
        names_list = []
        for i in orders_text:
            temp = i.split('|')
            if temp[1] not in names_list:
                names_list.append(temp[1])
        names_menu = types.InlineKeyboardMarkup()
        for i in names_list:
            if i == 'test':
                pass
            else:
                name_button = types.InlineKeyboardButton(text=i, callback_data=f'name_order: {i}')
                names_menu.add(name_button)
        bot.send_message(message.chat.id, 'Выбери чьё меню посмотреть', reply_markup=names_menu)


# Все меню одного человека
@bot.callback_query_handler(func=lambda msg: msg.data[:11] == 'name_order:')
def admin_person_order(callback):
    with open('Заказы.txt', 'r') as orders:
        text_orders = orders.readlines()
        name_callback = callback.data[12:]
        temp_list = []
        for i in range(len(text_orders)):
            text_orders[i] = text_orders[i].strip().split('|')
        for i in text_orders:
            if i[1] == name_callback:
                temp_list.append(i)
        temp_list = sorted(temp_list, key=lambda x: days_for_sort[x[0]])
        bot.send_message(callback.message.chat.id, name_callback)
        all_suma = 0
        for j in just_days:
            day_sum = 0
            counter = 0
            text_to_edit = bot.send_message(callback.message.chat.id, f'<b>{j}</b>', parse_mode='HTML')
            for i in temp_list:
                if i[0] == j:
                    counter += 1
                    bot.send_message(callback.message.chat.id, f'{counter}. {i[2]} <i>{i[3]}</i>', parse_mode='HTML')
                    day_sum += int(i[3])
                    all_suma += int(i[3])
            bot.edit_message_text(chat_id=callback.message.chat.id, message_id=text_to_edit.message_id, text=f'<b>{j}:</b>\nСумма заказа: <b>{day_sum}</b> ₽', parse_mode='HTML')
        bot.send_message(callback.message.chat.id, f'Сумма заказа за неделю: <b>{all_suma}</b> ₽', parse_mode='HTML')


# Файл с логами
@bot.message_handler(func=lambda msg: msg.text == 'Файл с логами')
def logs_file(message):
    with open('Логи.txt', 'rb') as logs_file:
        bot.send_document(message.chat.id, logs_file)


# Чеки
@bot.message_handler(func=lambda msg: msg.text == 'Чеки')
def admin_bills(message):
        bot.send_message(message.chat.id, 'В разработке')


# InLine-клавиатура Дни недели
@bot.message_handler(func=lambda msg: msg.text == 'Меню')
def menu(message):
    menu_week = types.InlineKeyboardMarkup(row_width=1)
    monday = types.InlineKeyboardButton(text='Понедельник', callback_data='Понедельник')
    tuesday = types.InlineKeyboardButton(text='Вторник', callback_data='Вторник')
    wednesday = types.InlineKeyboardButton(text='Среда', callback_data='Среда')
    thursday = types.InlineKeyboardButton(text='Четверг', callback_data='Четверг')
    friday = types.InlineKeyboardButton(text='Пятница', callback_data='Пятница')
    menu_week.add(monday, tuesday, wednesday, thursday, friday)
    menu_food = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    excel_button = types.KeyboardButton(text='Меню на неделю в Excel')
    back_button = types.KeyboardButton(text='Назад')
    menu_food.add(excel_button, back_button)
    bot.send_message(message.chat.id, 'Выбери день недели, чтобы посмотреть меню и сделать заказ', reply_markup=menu_food)
    bot.send_message(message.chat.id, 'Дни недели:', reply_markup=menu_week)


# Файл с excel меню
@bot.message_handler(func=lambda msg: msg.text == 'Меню на неделю в Excel')
def menu_excel(message):
    with open('Меню.xlsx', 'rb') as excel:
        bot.send_document(message.chat.id, excel)


# Inline-клавиатура с выбором еды
@bot.callback_query_handler(func=lambda msg: msg.data != 'Назад' and msg.data[:3] != 'del' and msg.data[:11] != 'name_order:')
def callback_all(callback):


    # Реакции на выбор блюда
    if callback.data not in ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница']:
        cell = callback.data.split()
        if len(cell) < 5:
            cell.append('_')
        temp_num = file_list.cell(row=int(cell[0]), column=int(cell[1])).value
        if temp_num == None:
            file_list.cell(row=int(cell[0]), column=int(cell[1])).value = 0
        file_list.cell(row=int(cell[0]), column=int(cell[1])).value += 1
        temp_num = file_list.cell(row=int(cell[0]), column=int(cell[1])).value
        file.save('Меню.xlsx')
        with open('Заказы.txt', 'r+') as orders_txt:
            user_food_name = cell[3] + ' ' + cell[4]
            food_name = file_list.cell(row=int(cell[0]), column=(int(cell[1])) - 1).value
            food_price = file_list.cell(row=int(cell[0]), column=(int(cell[1])) + 2).value
            id_food_name = id_generator(cell[3])
            new_food = f'{cell[2]}|{user_food_name}|{food_name}|{food_price}|{id_food_name}\n'
            orders_txt.seek(0, 2)
            orders_txt.write(new_food)
            orders_txt.seek(0)
            file_text = orders_txt.readlines()
            print(f'Сделан заказ: {cell[2]} {user_food_name} {food_name} {food_price} {id_food_name}')
            with open('Логи.txt', 'a') as logs:
                time = datetime.datetime.now()
                logs.write(f'{str(time)[:19]} Сделан заказ: {cell[2]} | {user_food_name} | {food_name} | {food_price} | {id_food_name}\n')
        suff_day = days_suff[cell[2]]
        bot.send_message(callback.message.chat.id, f'Твой заказ: "{food_name}" на <b>{suff_day}</b>', parse_mode='HTML')
    else:
        week_days = {'Понедельник': monday_menu, 'Вторник': tuesday_menu, 'Среда': wednesday_menu, 'Четверг': thursday_menu, 'Пятница': friday_menu}
        menu_for_the_day = types.InlineKeyboardMarkup()
        counter = 0
        callback_day = callback.data
        user_firstname = translit_name(callback.from_user.first_name)
        day = week_days[callback.data]
        button_order = types.InlineKeyboardButton(text=None, callback_data=None)
        for i in day:
            id_name_food = id_generator(user_firstname)
            user_name = translit_name(callback.from_user.first_name)
            user_lastname = translit_name(callback.from_user.last_name)
            temp_text = user_name + user_lastname
            button_order = types.InlineKeyboardButton(text=f'{i} ({day[i][0]} руб.)', callback_data=f'{day[i][1]} {day[i][2]} {callback_day} {user_name} {user_lastname}')
            menu_for_the_day.add(button_order)
            counter += 1
        button_back = types.InlineKeyboardButton(text='Выбрать другой день', callback_data='Назад')
        menu_for_the_day.add(button_back)
        bot.edit_message_text(chat_id=callback.message.chat.id, message_id=callback.message.id, text=callback_day, reply_markup=menu_for_the_day)


# Кнопка Назад
@bot.callback_query_handler(func=lambda msg: msg.data == 'Назад')
def inline_back(callback):
    menu_week = types.InlineKeyboardMarkup(row_width=1)
    monday = types.InlineKeyboardButton(text='Понедельник', callback_data='Понедельник')
    tuesday = types.InlineKeyboardButton(text='Вторник', callback_data='Вторник')
    wednesday = types.InlineKeyboardButton(text='Среда', callback_data='Среда')
    thursday = types.InlineKeyboardButton(text='Четверг', callback_data='Четверг')
    friday = types.InlineKeyboardButton(text='Пятница', callback_data='Пятница')
    menu_week.add(monday, tuesday, wednesday, thursday, friday)
    bot.edit_message_text(chat_id=callback.message.chat.id, message_id=callback.message.id, text='Выбери день недели', reply_markup=menu_week)


# Кнопка Оплаты
@bot.message_handler(func=lambda msg: msg.text == 'Оплатить заказ')
def pay_order(message):
    bot.send_message(message.chat.id, 'Вот ссылка для оплаты твоего заказа:\n'
                                      'https://www.tinkoff.ru/rm/savelova.tatyana1/wCseS42762\n'
                                      'После оплаты необходимо отправить боту фото с чеком\n')



#Кнопка Мой заказ
@bot.message_handler(func=lambda msg: msg.text == 'Мой заказ')
def my_order(message):
    order_menu = types.ReplyKeyboardMarkup(row_width=3, resize_keyboard=True)
    week_menu_button = types.KeyboardButton(text='Неделя')
    monday_button = types.KeyboardButton(text='Понедельник')
    tuesday_button = types.KeyboardButton(text='Вторник')
    wednesday_button = types.KeyboardButton(text='Среда')
    thursday_button = types.KeyboardButton(text='Четверг')
    friday_button = types.KeyboardButton(text='Пятница')
    back_button = types.KeyboardButton(text='Назад')
    delete_from_order = types.KeyboardButton(text='Удалить позиции из заказа')
    order_menu.add(monday_button, tuesday_button, wednesday_button)
    order_menu.add(tuesday_button, friday_button, week_menu_button)
    order_menu.add(delete_from_order)
    order_menu.add(back_button)
    bot.send_message(message.chat.id,'Выбери один из пунктов меню ⬇', reply_markup=order_menu)



# Кнопка Посмотреть заказ за неделю
@bot.message_handler(func=lambda msg: msg.text == 'Неделя')
def my_order_week(message):
    with open('Заказы.txt', 'r') as orders:
        name = message.from_user.first_name
        if message.from_user.last_name == None:
            lastname = '_'
        else:
            lastname = message.from_user.last_name
        name = translit_name(name)
        lastname = translit_name(lastname)
        user_fullname = name + ' ' + lastname
        text_orders = orders.readlines()
        all_suma = 0
        for j in range(5):
            day = just_days[j]
            text_to_edit = bot.send_message(message.chat.id, f'<b>{day}:</b>', parse_mode='HTML')
            suma_order = 0
            count = 1
            for i in text_orders:
                order_message = i.split('|')
                if order_message[0] == day and order_message[1] == user_fullname:
                    bot.send_message(message.chat.id, f'{count}. {order_message[2]} <i>{order_message[3]}</i>', parse_mode='HTML')
                    suma_order += int(order_message[3])
                    all_suma += int(order_message[3])
                    count += 1
            if suma_order != 0:
                bot.edit_message_text(chat_id=message.chat.id, message_id=text_to_edit.message_id, text = f'<b>{day}:</b>\n<u>Сумма заказа: {suma_order}</u> ₽', parse_mode='HTML')
            else:
                bot.edit_message_text(chat_id=message.chat.id, message_id=text_to_edit.message_id, text=f'<b>{day}:</b>\nНичего не заказано', parse_mode='HTML')
    bot.send_message(message.chat.id, f'<b>Итого:</b> <i>{all_suma}</i> ₽', parse_mode='HTML')


# Заказы по дням недели
@bot.message_handler(func=lambda msg: msg.text in just_days)
def day_order(message):
    with open('Заказы.txt', 'r') as orders:
        name = message.from_user.first_name
        if message.from_user.last_name == None:
            lastname = '_'
        else:
            lastname = message.from_user.last_name
        name = translit_name(name)
        lastname = translit_name(lastname)
        user_fullname = name + ' ' + lastname
        text_orders = orders.readlines()
        day = message.text
        text_to_edit = bot.send_message(message.chat.id, f'<b>{day}:</b>', parse_mode='HTML')
        suma_order = 0
        count = 1
        for i in text_orders:
            order_message = i.split('|')
            if order_message[0] == day and order_message[1] == user_fullname:
                bot.send_message(message.chat.id, f'{count}. {order_message[2]} <i>{order_message[3]}</i>', parse_mode='HTML')
                suma_order += int(order_message[3])
                count += 1
        if suma_order != 0:
            bot.edit_message_text(chat_id=message.chat.id, message_id=text_to_edit.message_id, text=f'<b>{day}:</b>\n<u>Сумма заказа: {suma_order}</u> ₽', parse_mode='HTML')
        else:
            bot.edit_message_text(chat_id=message.chat.id, message_id=text_to_edit.message_id, text=f'<b>{day}:</b>\nНичего не заказано', parse_mode='HTML')


# Кнопка Удалить позиции из заказа
@bot.message_handler(func=lambda msg: msg.text == 'Удалить позиции из заказа')
def delete_order_buttons(message):
    button_delete = types.InlineKeyboardButton(text=None, callback_data=None)
    with open('Заказы.txt') as order:
        new_temp = []
        text_order = order.readlines()
        menu_for_delete = types.InlineKeyboardMarkup()
        counter = 0
        for i in text_order:
            if len(i) < 3:
                continue
            else:
                user_firstname = message.from_user.first_name
                user_firstname = translit_name(user_firstname)
                button_order = types.InlineKeyboardButton(text=None, callback_data=None)
                if message.from_user.last_name == None:
                    user_lastname = '_'
                else:
                    user_lastname = message.from_user.last_name
                    user_lastname = translit_name(user_lastname)
                user_fool_name = user_firstname + ' ' + user_lastname
                name_of_order = i
                temp = i.strip().split('|')
                new_temp.append(temp)
        new_temp = sorted(new_temp, key = lambda x: days_for_sort[x[0]])
        for temp in new_temp:
            if user_fool_name in temp:
                price = temp[3]
                day = temp[0]
                id_food_name = temp[-1]
                name_food = temp[2]
                text_for_button = ' '.join(temp)
                button_delete = types.InlineKeyboardButton(text=f'{day}: {name_food} ({price} руб.)', callback_data=f'del{temp[-1]}')
                menu_for_delete.add(button_delete)
    bot.send_message(message.chat.id, 'Выбери, что хочешь удалить', reply_markup=menu_for_delete)


@bot.callback_query_handler(func=lambda msg: msg.data[:3] == 'del')
def delete_order(callback):
    id_for_delete = callback.data[3:]
    with open('Заказы.txt', 'r') as orders_file:
        orders_text = orders_file.readlines()
        temp = list(map(lambda x: x.split('|'), orders_text))
        new_orders = []
    with open('Заказы.txt', 'w') as orders_file:
        for i in temp:
            if id_for_delete == i[-1].strip():
                bot.send_message(callback.message.chat.id, f'{i[2]} на <b>{days_suff[i[0]]}</b> удалено', parse_mode='HTML')
                print(f'Удален заказ {" ".join(i)}')
                with open('Логи.txt', 'a') as logs:
                    time = datetime.datetime.now()
                    logs.write(f'{str(time)[:19]} Удален заказ: {" | ".join(i)}')
                continue
            else:
                new_orders.append(i)
        for i in new_orders:
            new_text = '|'.join(i)
            orders_file.write(f'{new_text}\n')


@bot.message_handler(content_types=['sticker'])
def send_sticker(message):
    sticker_id = message.sticker.file_id
    if sticker_id not in stickers_id:
        stickers_id.append(sticker_id)
    bot.send_sticker(message.chat.id, random.choice(stickers_id))


# Оплата заказа
@bot.message_handler(content_types=['photo'])
def send_to_operators(message):
    for i in operators_id:
        bot.forward_message(i, message.chat.id, message.message_id)
    menu_pay_day = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    monday_pay = types.KeyboardButton('Оплата за понедельник')
    tuesday_pay = types.KeyboardButton('Оплата за вторник')
    wednesday_pay = types.KeyboardButton('Оплата за среду')
    thursday_pay = types.KeyboardButton('Оплата за четверг')
    friday_pay = types.KeyboardButton('Оплата за пятницу')
    back_button = types.KeyboardButton('Назад')
    menu_pay_day.add(monday_pay, tuesday_pay, wednesday_pay, thursday_pay, friday_pay, back_button)
    bot.send_message(message.chat.id, 'Выбери за какой день была оплата ⬇', reply_markup=menu_pay_day)
    bill_list.append(message.message_id)


@bot.message_handler(func=lambda msg: msg.text in ['Оплата за понедельник', 'Оплата за вторник', 'Оплата за среду', 'Оплата за четверг', 'Оплата за пятницу'])
def pay_day(message):
    for i in operators_id:
        bot.forward_message(i, message.chat.id, message.message_id)


@bot.message_handler(commands=['id'])
def command_id(message):
    bot.send_message(message.chat.id, f'Name: {message.from_user.first_name}\n'
                                      f'Surname: {message.from_user.last_name}\n'
                                      f'Nickname: {message.from_user.username}\n'
                                      f'ID: {message.from_user.id}')


bot.send_message(198384562, 'Бот готов к работе')
bot.polling()

